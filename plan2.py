# coding=UTF-8
from tkinter import filedialog
from tkinter import *
from openpyxl.styles import Border, Side
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import openpyxl
import urllib
import urllib.request
from urllib import parse
import re
import datetime
from bs4 import BeautifulSoup
'''
    2021.3 增加界面GUI设计
    设置主菜单，打开文件

'''
from PyQt5.QtWidgets import (QMainWindow,QTextEdit,QAction,QFileDialog,QApplication)
from PyQt5.QtGui import QIcon
import sys

# 从字符串得到 年 月，如：2020-09 返回元组：{"2020','-','09'}
def getyearmonth(datestr):
    datepattern = re.compile(r'(?P<year>\d{4})(-|/)(?P<month>\d{2})')  # yyyy-mm年月格式
    result = re.match(datepattern, datestr)
    return result.groups()


# 计算月份差,返回 int 型值
def datediff(datestr1, datestr2):
    if datestr1 == datestr2:
        return 0
    else:
        date1 = getyearmonth(datestr1)
        date2 = getyearmonth(datestr2)
        if date1[0] == date2[0]:
            return int(date1[2]) - int(date2[2])
        else:
            return (int(date1[0]) - int(date2[0])) * 12 + int(date1[2]) - int(date2[2])


# 计算下次段修日期
def getxcdxrq(ycdx, bzy):
    year = int(bzy) // 12
    month = int(bzy) % 12
    ycdxny = getyearmonth(ycdx)
    year = year + int(ycdxny[0])
    month = month + int(ycdxny[2])
    if month > 12:
        month = month % 12
        year = year + 1
    if month < 10:
        return str(year) + '-0' + str(month)
    else:
        return str(year) + '-' + str(month)


# 计算本次做厂修的下次厂修日期，当前年月+厂修周期
def getxcxrq(cxzq):
    year = int(cxzq) // 12
    month = int(cxzq) % 12
    yccxny = getyearmonth(nowdatestr)
    year = year + int(yccxny[0])
    month = month + int(yccxny[2])
    if month > 12:
        month = month % 12
        year = year + 1
    if month < 10:
        return str(year) + '-0' + str(month)
    else:
        return str(year) + '-' + str(month)


# 判断是否加强段修
# 判断条件:2020 2021年，C70 P70 X70型车首次厂修到期的做一次加强段修
# 判断标准：前次厂修=制造年月  下次厂修-前次厂修>=8
def ifjq(cx, qccx, zzrq, xcdx):
    if str(nowyear) == '2020' or str(nowyear) == '2021':
        cxpattern = re.compile(r'C70|P70|X70')
        cxresult = re.match(cxpattern, cx)
        if cxresult:
            qccxg = re.match(r'\d{4}', qccx)
            zzrqg = re.match(r'\d{4}', zzrq)
            qccxn = qccxg.group()
            zzrqn = zzrqg.group()
            if qccxn == zzrqn:
                xcdxg = re.match(r'\d{4}', xcdx)
                xcdxn = xcdxg.group()
                if int(xcdxn) - int(qccxn) >= 8:
                    return True
                else:
                    return False  # 原厂修未到期（不大于8年）
            else:
                return False  # 前次厂修<>制造年月（不是首次）
        else:
            return False  # 非C70 P70 X70
    else:
        return False  # 非2020 2021年


# 获取车辆基础信息
def getjsll(dataurl, ch, rq):
    bzyf = ''
    cxyq = ''
    for line, chehao in ch.items():
        sbrqdatestr = rq[line]
        info = []
        postdata = {'ch': chehao}
        headers = {'User-Agent': 'User-Agent:Mozilla/5.0'}
        data = urllib.parse.urlencode(postdata).encode('utf-8')
        req = urllib.request.Request(dataurl, data, headers=headers)
        response = urllib.request.urlopen(req)


        try:
            html = response.read().decode('utf-8')
        except:
            print("未连接到目标网络，请稍后再试")
        else:
            bs = BeautifulSoup(html, 'html.parser')
            chexing = bs.find("td", text='车型').next_sibling.next_sibling.string
            zzrq = bs.find("td", text='制造日期').next_sibling.next_sibling.string
            zzdw = bs.find("td", text='制造单位').next_sibling.next_sibling.string

            qccxrqstr = bs.find("td", text='厂修时间基准').next_sibling.next_sibling.string
            qccxpattern = re.compile(r'(?P<year>\d{4})(-|/)(?P<month>\d{2})')  # yyyy-mm年月格式
            qccxrqgroup = re.match(qccxpattern, qccxrqstr)
            qccxrq = qccxrqgroup.group()
            mcxc = bs.find("td", text='末次检修修程').next_sibling.next_sibling.string
            if mcxc == '厂修' or mcxc == '新造' or mcxc == '入段厂修':
                qccxdw = bs.find("td", text='末次检修单位').next_sibling.next_sibling.string
            else:
                qccxdw = bs.find("td", text='前次厂修单位').next_sibling.next_sibling.string

            # 新造车无未做过临修，前次厂修、前次段修信息为None,直接用新造时间单位填补
            if qccxrq is None:
                qccxrq = zzrq
                qccxdw = re.sub(r"\s+", "", zzdw)
            # 前次段修有关信息（保留不用）
            # qcdxrq = bs.find("td", text='前次段修日期').next_sibling.next_sibling.string
            # qcdxdw = bs.find("td", text='前次段修单位').next_sibling.next_sibling.string
            ycdxrq = bs.find("td", text='下次段修日期').next_sibling.next_sibling.string  # 本次施修前预测的下次段修日期
            cxzq = bs.find("td", text=re.compile('厂修周期')).next_sibling.next_sibling.string  # 厂修周期
            dxzq = bs.find("td", text=re.compile('段修周期')).next_sibling.next_sibling.string  # 段修周期

            # 读取厂修延期时间
            length = len(bs.find('td', text='天数').parent.next_sibling.next_sibling.contents)
            for i in range(1, length, 2):
                strtemp = bs.find('td', text='天数').parent.next_sibling.next_sibling.contents[i].string
                if strtemp is None:
                    cxyq = ''
                    break
                elif re.match(r'\d{4}-\d{2}-\d{2}', strtemp):
                    pass
                elif re.match(r'\d', strtemp):
                    cxyq = strtemp
                    break

            xccxrqstr = bs.find("td", text='下次厂修日期').next_sibling.next_sibling.string

            # 如未预测下次段修则是厂修到期,下次厂修为当前年月+厂修周期，下次段修为当前年月+段修周期，保证月份为段修周期
            # 如预测下次段修,下次厂修为读取的下次厂修日期（取年月），下次段修为当前年月+保证月份（读取的月份）
            if ycdxrq is None:  # 未预测下次段修说明厂修到期
                bzyf = dxzq  # 厂修到期的保证月份为段修周期
                xccxrq = getxcxrq(cxzq)  # 下次厂修为当前年月+厂修周期
                cxyq = ''   # 如厂修则延期月无用设为空
            else:  # 预测了下次段修
                nypattern = re.compile(r'(?P<year>\d{4})(-|/)(?P<month>\d{2})')  # 年-月格式
                xccxgroup = re.match(nypattern, xccxrqstr)
                xccxrq = xccxgroup.group()
                # 如果是加强段修
                if ifjq(chexing, qccxrq, zzrq, ycdxrq):
                    bzyf = str(datediff(xccxrq, sbrqdatestr))
                # 如果不是加强段修
                else:
                    bzyf = str(datediff(ycdxrq, sbrqdatestr) + int(dxzq))
            xcdxrq = getxcdxrq(sbrqdatestr, bzyf)

            # 将基本信息保存在列表中
            info.append(zzrq)
            info.append(qccxrq)
            info.append(qccxdw)
            # info.append(qcdxrq)  # 前次段修日期（保留不用）
            # info.append(qcdxdw)  # 前次段修单位（保留不用）
            info.append(xccxrq)
            info.append(xcdxrq)
            info.append(cxyq)
            info.append(bzyf)
            if ycdxrq is None:
                info.append('')
            else:
                info.append(ycdxrq)

            # 将列表中的基本信息添加到字典（基础信息）中
            infoitems = {line: info}
            jsll.update(infoitems)



def rwexcell(filename):
    wb = openpyxl.load_workbook(filename)
    allSheets = wb.sheetnames
    ws = wb.get_sheet_by_name(allSheets[0])

    chehao = {}  # 得到车号列表
    sbrq = {}  # 上报日期列表
    chehaocol = 0  # 车号数据所在单元格列数
    chehaorow = 0  # 车号数据所在单元格行数

    # 查找到车号所在的列，及车号开始的行
    for i in range(1, ws.max_column):
        for j in range(1, ws.max_row):
            if ws.cell(column=i, row=j).value == u'车号':
                chehaocol = i
                chehaorow = j
                break

    # 查找到前次厂段修信息所在单元格信息
    qccxsjcol = 0  # 前次厂修时间列
    qccxdwcol = 0  # 前次厂修单位列
    qcdxsjcol = 0  # 前次段修时间列
    qcdxdwcol = 0  # 前次段修单位列
    zznycol = 0  # 制造年月列号
    bzyfcol = 0  # 保证月份列号
    sbrqcol = 0  # 上报日期列号
    xccxcol = 0  # 下次厂修列号（要添加的列）
    xcdxcol = 0  # 下次段修列号（要添加的列）
    cxyqcol = 0  # 厂修延期列号（要添加的列）


    # 查找前次厂修、前次段修、制造年月、保证月份所在列的列号
    for i in range(1, ws.max_column):
        for j in range(1, ws.max_row):
            if ws.cell(column=i, row=j).value == u'前次厂修':
                qccxsjcol = i
                qccxdwcol = i + 1
            if ws.cell(column=i, row=j).value == u'前次段修':
                qcdxsjcol = i
                qcdxdwcol = i + 1
            if ws.cell(column=i, row=j).value == u'制造年月':
                zznycol = i
            if ws.cell(column=i, row=j).value == u'保证月份':
                bzyfcol = i
            if ws.cell(column=i, row=j).value == u'下次厂修':
                xccxcol = i
            if ws.cell(column=i, row=j).value == u'下次段修':
                xcdxcol = i
            if ws.cell(column=i, row=j).value == u'历史延期':
                cxyqcol = i
            if ws.cell(column=i, row=j).value == u'上报HMIS时间':
                sbrqcol = i

    # 识别出车号，并保存在车号列表中
    rule = re.compile(r'(^[1-9](\d){6})')
    for i in range(chehaorow + 2, ws.max_row):
        if rule.match(str(ws.cell(column=chehaocol, row=i).value)):
            chehao[i] = ws.cell(column=chehaocol, row=i).value
            if ws.cell(column=sbrqcol, row=i).value == '':
                sbrq[i] = nowdatestr
            else:
                rq = ws.cell(column=sbrqcol, row=i).value
                if rq.month < 10:
                    sbrq[i] = str(rq.year) + '-0' + str(rq.month)
                else:
                    sbrq[i] = str(rq.year) + '-' + str(rq.month)

    # print(chehao,sbrq)

    # 获取基础信息开始
    getjsll(url, chehao, sbrq)
    # print(jsll)

    # 处理EXCEL表格开始,开始对EXCEL表格进行写操作
    # 简化表格：去掉表格中多余 的'转、型、阀'等字
    for i in range(chehaorow + 1, ws.max_row):
        for j in range(ws.max_column):
            if "型" in str(ws.cell(i + 1, j + 1).value):
                temp = str(ws.cell(i + 1, j + 1).value).replace('型', '')
                ws.cell(i + 1, j + 1).value = temp
            elif "阀" in str(ws.cell(i + 1, j + 1).value):
                temp = str(ws.cell(i + 1, j + 1).value).replace('阀', '')
                ws.cell(i + 1, j + 1).value = temp
            elif "修" in str(ws.cell(i + 1, j + 1).value):
                temp = str(ws.cell(i + 1, j + 1).value).replace('修', '')
                ws.cell(i + 1, j + 1).value = temp
            elif "铁" in str(ws.cell(i + 1, j + 1).value):
                temp = str(ws.cell(i + 1, j + 1).value).replace('铁', '')
                ws.cell(i + 1, j + 1).value = temp
    # 把前次厂修名称改为'厂修基准'
    ws.cell(chehaorow, qccxsjcol).value = '厂修基准'

    # 在前次段修单位列后插入列：本次段修到期时间，下次厂修、下次段修日期，历史延期，共4列
    if xccxcol == 0:  # 如果不存在下次厂修、下次段修、延期月三列
        ws.insert_cols(qcdxdwcol + 1, 4)
        # 相应的修改制造年月、保证月份列数
        ycdxcol = qcdxdwcol + 1  # 本次段修到期时间
        xccxcol = qcdxdwcol + 2  # 下次厂修到期
        xcdxcol = qcdxdwcol + 3  # 下次段修到期
        cxyqcol = qcdxdwcol + 4  # 延期月份
        zznycol = zznycol + 4  # 制造年月
        bzyfcol = bzyfcol + 4  # 保证月份
        # 添加边框
        ft = Font(name='宋体', size=12)  # 设置字体及大小
        align = Alignment(horizontal='center', vertical='top', wrap_text=True)  # 对齐方式,自动换行
        border_set = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                            bottom=Side(style='thin'))
        for columns in range(ycdxcol, cxyqcol + 1):
            for rows in range(chehaorow, ws.max_row):
                ws.cell(rows, columns).border = border_set
                ws.cell(rows, columns).font = ft
                ws.cell(rows, columns).alignment = align
        # 设置表头信息
        ft2 = Font(name='黑体', size=12)  # 设置字体及大小
        ws.cell(chehaorow, ycdxcol).font = ft2
        ws.cell(chehaorow, ycdxcol).value = "本次段修到期"
        ws.cell(chehaorow, xccxcol).font = ft2
        ws.cell(chehaorow, xccxcol).value = "下次厂修"
        ws.cell(chehaorow, xcdxcol).font = ft2
        ws.cell(chehaorow, xcdxcol).value = "下次段修"
        ws.cell(chehaorow, cxyqcol).font = ft2
        ws.cell(chehaorow, cxyqcol).value = "历史延期"

    # 将基础数据写入单元格
    for row in jsll.keys():
        print("%s :上报 %s: %s" % (chehao[row], sbrq[row], jsll[row]))
        jsinfo = jsll[row]
        ws.cell(row, zznycol).value = jsinfo[0]  # 制造年月
        ws.cell(row, qccxsjcol).value = jsinfo[1]  # 前次厂修时间
        ws.cell(row, qccxdwcol).value = jsinfo[2]  # 前次厂修单位
        # ws.cell(row, qcdxsjcol).value = jsinfo[3]   # 前次段修时间
        # ws.cell(row, qcdxdwcol).value = jsinfo[4]   # 前次段修单位
        ws.cell(row, xccxcol).value = jsinfo[3]  # 下次厂修时间
        ws.cell(row, xcdxcol).value = jsinfo[4]  # 下次段修时间
        ws.cell(row, cxyqcol).value = jsinfo[5]  # 延期月份
        ws.cell(row, bzyfcol).value = jsinfo[6]  # 保证月份
        ws.cell(row, ycdxcol).value = jsinfo[7]  # 本次段修到期

    # 隐藏上次段修有关信息（保留）
    ws.column_dimensions[get_column_letter(qcdxsjcol)].visible = False
    ws.column_dimensions[get_column_letter(qcdxdwcol)].visible = False

    # 写数据到EXCELL表格
    try:
        wb.save(fn)
    except IOError:
        print("错误：文件无法写入，请检查是否当前打开，请关闭后再处理！")
    else:
        print('---处理完毕---')

class Example(QMainWindow):

    def __init__(self) :
        super().__init__()

        self.initUI()

    def initUI(self):

        self.textEdit = QTextEdit()
        self.setCentralWidget(self.textEdit)
        self.statusBar()

        openFile = QAction(QIcon('open.png'),'打开文件',self)
        openFile.setShortcut('Ctrl+O')
        openFile.setStatusTip('打开xlsx文件')
        openFile.triggered.connect(self.showDialog)

        menubar = self.menuBar()
        fileMenu = menubar.addMenu('&File')
        fileMenu.addAction(openFile)

        self.setGeometry(300,300,350,300)
        self.setWindowTitle('File dialog')
        self.show()

    def showDialog(self):

        fname = QFileDialog.getOpenFileName(self,'Open file','/home')

        if fname[0]:
            f = open(fname[0],'r')

            rwexcell(f)



if __name__ == '__main__':
    # 基础数据URL地址http://10.3.102.100:8084
    url = 'http://10.3.102.100:8084/hcxt/reportFiles/ch_info_query/jsll2.jsp?file=jsll_chzhquery.raq'

    # 基础信息字典变量：jsll
    jsll = {}

    # 从系统获取当前日期的年月，并保存为"2020-09"格式的字符串变量: nowdatestr
    nowyear = datetime.datetime.now().year
    nowmonth = datetime.datetime.now().month
    if nowmonth < 10:
        nowdatestr = str(nowyear) + '-0' + str(nowmonth)
    else:
        nowdatestr = str(nowyear) + '-' + str(nowmonth)

    app = QApplication(sys.argv)
    ex = Example()
    sys.exit(app.exec_())