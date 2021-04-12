from tkinter import filedialog
from tkinter import *
import openpyxl
import re
from bs4 import BeautifulSoup

root = Tk()
root.filename = filedialog.askopenfilename(initialdir="/", title="Select file", filetypes=(("excel files","*.xlsx"), ("all files","*.*")))
print(root.filename)

fn = root.filename
wb = openpyxl.load_workbook(fn)
allSheets = wb.sheetnames
ws = wb.get_sheet_by_name(allSheets[0])
print("所有工作表", allSheets)
print("当前工作表", ws.title)
print("当前工作表栏数：", ws.max_column)
print("当前工作表行数：", ws.max_row)

chehao = {}  #得到车号列表
chehaocol = 0  #车号数据所在单元格列数
chehaorow = 0  #车号数据所在单元格行数


#查找到车号所在的列，及车号开始的行
for i in range(1, ws.max_column) :
    for j in range(1, ws.max_row) :
        if ws.cell(column=i, row=j).value == '车号':
            chehaocol = i
            chehaorow = j
            break

#查找到前次厂段修信息所在单元格信息
qccxsjcol = 0
qccxdwcol = 0
qcdxsjcol = 0
qcdxdwcol = 0
for i in range(1, ws.max_column) :
    for j in range(1, ws.max_row) :
        if ws.cell(column=i, row=j).value == '前次厂修':
            qccxsjcol = i
            qccxdwcol = i+1
        if ws.cell(column=i, row=j).value == '前次段修':
            qcdxsjcol = i
            qcdxdwcol = i+1


print("车号在：%s列 %s行" % (chehaocol, chehaorow))
print("前次厂修时间在：%s列 前次厂修单位%s 列" % (qccxsjcol, qccxdwcol))
print("前次段修时间在：%s列 前次段修单位%s 列" % (qcdxsjcol, qcdxdwcol))

#识别出车号，并保存在车号列表中
rule = re.compile(r'\d{7}')

for i in range(chehaorow,ws.max_row):
    if rule.match(str(ws.cell(column=chehaocol, row=i).value)):
        chehao[i] = ws.cell(column=chehaocol, row=i).value
print(chehao, end=' ')